# import pandas as pd
 
# # Create months list for 36 months starting Jan 2026
# months = pd.date_range(start="2026-01-01", periods=36, freq='MS').strftime("%b %Y")
 
# # Initialize data lists
# salary = []
# expenses = []
# sip = []
# insurance = []
# emergency_fund = []
# marriage_fund = []
# notes = []
 
# # Initial values
# current_salary = 32000
# increment_date = pd.Timestamp("2027-07-01")
# emergency_fund_total = 85000  # FD 50k + Liquid 35k
# emergency_fund_target = 100000
# emergency_topup_done = False
 
# for month in pd.date_range(start="2026-01-01", periods=36, freq='MS'):
#     # Salary progression
#     if month < pd.Timestamp("2026-03-01"):
#         sal = 32000
#     elif month < pd.Timestamp("2026-08-01"):
#         sal = 45000
#     else:
#         sal = 50000
 
#     # Apply 7% increment from 1 July 2027
#     if month >= increment_date:
#         months_since_inc = (month.year - 2027) * 12 + (month.month - 7)
#         sal = int(50000 * (1 + 0.07) ** (months_since_inc // 12 + 1))
 
#     salary.append(sal)
 
#     # Expenses (slightly increase with salary)
#     if sal <= 32000:
#         expenses.append(18000)
#     elif sal <= 45000:
#         expenses.append(20000)
#     else:
#         expenses.append(22000)
 
#     # SIP (long-term) progression
#     if sal <= 32000:
#         sip.append(5000)
#     elif sal <= 45000:
#         sip.append(5000)
#     else:
#         sip.append(5000 if month < increment_date else 8000)
 
#     # Insurance (Health + Term + Parents)
#     insurance.append(4000 if month.month <= 2 else 0)  # assuming paid in first 2 months
 
#     # Emergency Fund progress (top-up until target reached)
#     if not emergency_topup_done:
#         topup = min(emergency_fund_target - emergency_fund_total, 15000)
#         emergency_fund_total += topup
#         emergency_fund.append(topup)
#         if emergency_fund_total >= emergency_fund_target:
#             emergency_topup_done = True
#     else:
#         emergency_fund.append(0)
 
#     # Marriage Fund allocation
#     if sal <= 32000:
#         marriage_fund.append(5000)
#     elif sal <= 45000:
#         marriage_fund.append(10000)
#     else:
#         marriage_fund.append(15000 if month < increment_date else 18000)
 
#     # Notes
#     note = ""
#     if month == pd.Timestamp("2026-01-01"):
#         note = "Start insurance; top-up emergency fund"
#     elif month == pd.Timestamp("2026-03-01"):
#         note = "Salary hike to 45k; increase marriage fund"
#     elif month == pd.Timestamp("2026-08-01"):
#         note = "Salary increase to 50k; aggressive marriage fund allocation"
#     elif month == increment_date:
#         note = "Apply 7% annual increment"
#     notes.append(note)
 
# # Create DataFrame
# df = pd.DataFrame({
#     "Month": months,
#     "Salary (₹)": salary,
#     "Expenses (₹)": expenses,
#     "SIP (₹)": sip,
#     "Insurance (₹)": insurance,
#     "Emergency Fund Top-up (₹)": emergency_fund,
#     "Marriage Fund (₹)": marriage_fund,
#     "Notes": notes
# })
 
# # Save to Excel
# df.to_excel("36_month_financial_roadmap.xlsx", index=False)
# print("Excel file '36_month_financial_roadmap.xlsx' created successfully!")

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
 
# Parameters
months = pd.date_range(start="2026-01-01", periods=36, freq='MS').strftime("%b %Y")
fd_amount = 50000
liquid_amount = 35000
emergency_target = 100000
marriage_target = 1000000  # example target 10 lakh
base_sip = 5000
insurance_first_month = 4000
increment_pct = 0.07
increment_start_month = "Jul 2027"
 
# Create initial DataFrame
df = pd.DataFrame({
    "Month": months,
    "Salary (₹)": [32000]*2 + [45000]*5 + [50000]*29,  # initial salaries
    "Expenses (₹)": 0,
    "SIP (₹)": 0,
    "Insurance (₹)": 0,
    "Emergency Top-up (₹)": 0,
    "Emergency Total (₹)": 0,
    "Marriage Fund Contribution (₹)": 0,
    "Marriage Fund Total (₹)": 0,
    "Notes": ""
})
 
# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Financial Roadmap"
 
# Write headers
for col_num, column_title in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column_title)
    cell.font = Font(bold=True)
 
# Write months and placeholders
for r_idx, month in enumerate(months, start=2):
    ws.cell(row=r_idx, column=1, value=month)
 
# Setup parameters area (top of sheet)
param_row = 1
ws['L1'] = "FD Amount"
ws['M1'] = fd_amount
ws['L2'] = "Liquid Amount"
ws['M2'] = liquid_amount
ws['L3'] = "Emergency Target"
ws['M3'] = emergency_target
ws['L4'] = "Marriage Target"
ws['M4'] = marriage_target
ws['L5'] = "Base SIP"
ws['M5'] = base_sip
ws['L6'] = "Insurance First Month"
ws['M6'] = insurance_first_month
ws['L7'] = "Increment %"
ws['M7'] = increment_pct
ws['L8'] = "Increment Start Month"
ws['M8'] = increment_start_month
 
# Fill Excel with formulas
for row in range(2, 38):  # 36 months
    # Salary formula: apply increment if month >= increment start
    salary_cell = ws.cell(row=row, column=2)
    month_cell = ws.cell(row=row, column=1)
    if row < 8:  # before increment
        salary_cell.value = f"=IF({month_cell.coordinate}<\"{increment_start_month}\",{df.loc[row-2,'Salary (₹)']},{df.loc[row-2,'Salary (₹)']})"
    else:
        salary_cell.value = f"={df.loc[row-2,'Salary (₹)']}*(1+$M$7)"
 
    # Expenses: 18k initial, increase with salary
    expenses_cell = ws.cell(row=row, column=3)
    expenses_cell.value = f"=ROUND(0.4*{salary_cell.coordinate},0)"  # example: 40% of salary
 
    # SIP: base SIP, can adjust
    sip_cell = ws.cell(row=row, column=4)
    sip_cell.value = f"=$M$5"
 
    # Insurance: first month only
    insurance_cell = ws.cell(row=row, column=5)
    if row == 2:
        insurance_cell.value = f"=$M$6"
    else:
        insurance_cell.value = 0
 
    # Emergency Top-up formula
    emergency_topup_cell = ws.cell(row=row, column=6)
    if row == 2:
        emergency_topup_cell.value = f"=MAX(0,$M$3-($M$1+$M$2))"
    else:
        prev_total = ws.cell(row=row-1, column=7).coordinate
        emergency_topup_cell.value = f"=MAX(0,$M$3-{prev_total})"
 
    # Emergency Total formula
    emergency_total_cell = ws.cell(row=row, column=7)
    if row == 2:
        emergency_total_cell.value = f"=$M$1+$M$2+{emergency_topup_cell.coordinate}"
    else:
        prev_total = ws.cell(row=row-1, column=7).coordinate
        emergency_total_cell.value = f"={prev_total}+{emergency_topup_cell.coordinate}"
 
    # Marriage Fund Contribution formula (editable)
    marriage_contrib_cell = ws.cell(row=row, column=8)
    marriage_contrib_cell.value = f"=MIN(20000,{salary_cell.coordinate}-{expenses_cell.coordinate}-{sip_cell.coordinate}-{insurance_cell.coordinate})"
 
    # Marriage Fund Total formula
    marriage_total_cell = ws.cell(row=row, column=9)
    if row == 2:
        marriage_total_cell.value = f"={marriage_contrib_cell.coordinate}"
    else:
        prev_total = ws.cell(row=row-1, column=9).coordinate
        marriage_total_cell.value = f"={prev_total}+{marriage_contrib_cell.coordinate}"
 
    # Notes
    notes_cell = ws.cell(row=row, column=10)
    if row == 2:
        notes_cell.value = "Start insurance; top-up emergency fund"
    elif row == 3:
        notes_cell.value = "Salary hike; increase marriage fund"
    elif row == 8:
        notes_cell.value = "Salary increase to 50k; aggressive marriage fund"
    elif row == 19:
        notes_cell.value = "Apply 7% annual increment"
 
# Save file
wb.save("36_month_financial_roadmap_with_formulas.xlsx")
print("Excel sheet '36_month_financial_roadmap_with_formulas.xlsx' created successfully!")