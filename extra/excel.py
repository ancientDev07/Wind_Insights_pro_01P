
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Alignment, Font

# # Create workbook and sheets
# wb = Workbook()
# assumptions = wb.active
# assumptions.title = "Assumptions"
# projection = wb.create_sheet("Projection")

# # -----------------------------
# # ASSUMPTIONS SHEET
# # -----------------------------
# assumptions_data = [
#     ("Parameter", "Value"),
#     ("Start Year", 2025),
#     ("End Year", 2050),
#     ("Your Start Age", 23),
#     ("Parents Elder Age", 59),
#     ("Parents Younger Age", 54),
#     ("Your Premium 2025", 14400),
#     ("Parents Premium 2025", 36000),
#     ("Your Inflation (<35)", 0.06),
#     ("Your Inflation (35–45)", 0.08),
#     ("Your Inflation (45+)", 0.10),
#     ("Parents Inflation (60+)", 0.10),  # If you later want <60, add a new row and adjust formulas below.
# ]

# for row in assumptions_data:
#     assumptions.append(row)

# # Style assumptions header
# assumptions["A1"].font = Font(bold=True)
# assumptions["B1"].font = Font(bold=True)

# # -----------------------------
# # PROJECTION SHEET HEADERS
# # -----------------------------
# projection_headers = [
#     "Year",
#     "Your Age",
#     "Your Inflation %",
#     "Your Premium (₹)",
#     "Parents Elder Age",
#     "Parents Younger Age",
#     "Parents Inflation %",
#     "Parents Premium (₹)",
#     "Total Premium (₹)"
# ]
# projection.append(projection_headers)

# # Style header
# for col_idx, header in enumerate(projection_headers, start=1):
#     cell = projection.cell(row=1, column=col_idx)
#     cell.font = Font(bold=True)
#     cell.alignment = Alignment(horizontal="center")

# # Convenience absolute references (safer if rows/cols shift)
# START_YEAR   = "$B$2"
# END_YEAR     = "$B$3"
# YOUR_START   = "$B$4"
# P_ELDER      = "$B$5"
# P_YOUNGER    = "$B$6"
# YOUR_PREM0   = "$B$7"
# P_PREM0      = "$B$8"
# Y_INF_LT35   = "$B$9"
# Y_INF_35_45  = "$B$10"
# Y_INF_45PLUS = "$B$11"
# P_INF_60PLUS = "$B$12"

# # -----------------------------
# # FIRST YEAR FORMULAS (2025)
# # -----------------------------
# # Row 2 = first projection year
# projection["A2"] = f"=Assumptions!{START_YEAR}"      # Year
# projection["B2"] = f"=Assumptions!{YOUR_START}"      # Your Age

# # Your Inflation % based on age brackets
# projection["C2"] = (
#     f"=IF(B2<35,Assumptions!{Y_INF_LT35},"
#     f"IF(B2<45,Assumptions!{Y_INF_35_45},Assumptions!{Y_INF_45PLUS}))"
# )

# # Your Premium (base year)
# projection["D2"] = f"=Assumptions!{YOUR_PREM0}"

# # Parents Ages (numeric, separate columns)
# projection["E2"] = f"=Assumptions!{P_ELDER}"
# projection["F2"] = f"=Assumptions!{P_YOUNGER}"

# # Parents Inflation % (constant; if you add <60 later, change this to an IF)
# projection["G2"] = f"=Assumptions!{P_INF_60PLUS}"

# # Parents Premium (base year)
# projection["H2"] = f"=Assumptions!{P_PREM0}"

# # Total Premium
# projection["I2"] = "=D2+H2"

# # -----------------------------
# # FUTURE YEARS (AUTO FORMULAS)
# # -----------------------------
# # Fill rows until End Year (inclusive)
# start_row = 2
# # We'll build until the value in A(row) equals End Year
# # Year progression: row r holds year A(r) = A(r-1) + 1

# # Determine how many rows (years) we need based on numeric assumption
# start_year = assumptions["B2"].value
# end_year = assumptions["B3"].value
# num_years = (end_year - start_year) + 1

# for idx in range(1, num_years):  # idx=1 -> row3 (2026), ..., last -> 2050
#     row = start_row + idx
#     prev = row - 1

#     # Year and Age progression
#     projection[f"A{row}"] = f"=A{prev}+1"
#     projection[f"B{row}"] = f"=B{prev}+1"

#     # Your Inflation % (age bracket)
#     projection[f"C{row}"] = (
#         f"=IF(B{row}<35,Assumptions!{Y_INF_LT35},"
#         f"IF(B{row}<45,Assumptions!{Y_INF_35_45},Assumptions!{Y_INF_45PLUS}))"
#     )

#     # Your Premium with compound growth (rounded to nearest ₹)
#     projection[f"D{row}"] = f"=ROUND(D{prev}*(1+C{row}),0)"

#     # Parents Ages from start + year offset (no string hacks)
#     # offset = current_year - start_year = A(row) - Assumptions!StartYear
#     projection[f"E{row}"] = f"=Assumptions!{P_ELDER} + (A{row}-Assumptions!{START_YEAR})"
#     projection[f"F{row}"] = f"=Assumptions!{P_YOUNGER} + (A{row}-Assumptions!{START_YEAR})"

#     # Parents Inflation % (constant). If you add <60 later, use:
#     # =IF(MAX(E{row},F{row})>=60,Assumptions!$B$12,Assumptions!$B$13)
#     projection[f"G{row}"] = f"=Assumptions!{P_INF_60PLUS}"

#     # Parents Premium (compound growth; rounded)
#     projection[f"H{row}"] = f"=ROUND(H{prev}*(1+G{row}),0)"

#     # Total Premium
#     projection[f"I{row}"] = f"=D{row}+H{row}"

# # -----------------------------
# # FORMATTING
# # -----------------------------
# # Set column widths
# widths = {
#     "A": 8,   # Year
#     "B": 9,   # Your Age
#     "C": 14,  # Your Inflation %
#     "D": 16,  # Your Premium
#     "E": 17,  # Parents Elder Age
#     "F": 18,  # Parents Younger Age
#     "G": 18,  # Parents Inflation %
#     "H": 18,  # Parents Premium
#     "I": 16,  # Total Premium
# }
# for col_letter, w in widths.items():
#     projection.column_dimensions[col_letter].width = w

# # Number formats
# for r in range(2, 2 + num_years):
#     projection[f"C{r}"].number_format = "0.00%"
#     projection[f"D{r}"].number_format = '₹#,##0'
#     projection[f"E{r}"].number_format = '0'
#     projection[f"F{r}"].number_format = '0'
#     projection[f"G{r}"].number_format = "0.00%"
#     projection[f"H{r}"].number_format = '₹#,##0'
#     projection[f"I{r}"].number_format = '₹#,##0'

# # Freeze header row
# projection.freeze_panes = "A2"

# # -----------------------------
# # SAVE FILE
# # -----------------------------
# wb.save("Health_Insurance_Projection.xlsx")
# print("Excel file created successfully: Health_Insurance_Projection.xlsx")

from openpyxl import Workbook

# =========================
# CREATE WORKBOOK
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Insurance Projection"

# =========================
# INPUT SECTION
# =========================
ws["A1"] = "INPUTS"

ws["A3"] = "Start Year"
ws["B3"] = 2026

ws["A4"] = "End Year"
ws["B4"] = 2050

ws["A6"] = "Monthly Life Insurance Premium"
ws["B6"] = 1000

ws["A7"] = "Monthly Personal Medical Premium"
ws["B7"] = 1200

ws["A8"] = "Monthly Family Medical Premium"
ws["B8"] = 3000

ws["A10"] = "Life Insurance Inflation"
ws["B10"] = 0.00

ws["A11"] = "Medical Insurance Inflation"
ws["B11"] = 0.12

# =========================
# TABLE HEADERS
# =========================
start_row = 14

headers = [
    "Year",
    "Annual Life Insurance",
    "Annual Personal Medical Insurance",
    "Annual Family Medical Insurance",
    "Total Annual Insurance Cost",
    "Cumulative Insurance Cost"
]

for col, header in enumerate(headers, start=1):
    ws.cell(row=start_row, column=col, value=header)

# =========================
# FIRST YEAR FORMULAS
# =========================
ws["A15"] = "=B3"                        # Year
ws["B15"] = "=B6*12"                     # Life Insurance
ws["C15"] = "=B7*12"                     # Personal Medical
ws["D15"] = "=B8*12"                     # Family Medical
ws["E15"] = "=SUM(B15:D15)"              # Total
ws["F15"] = "=E15"                       # Cumulative

# =========================
# SUBSEQUENT YEARS (FORMULAS)
# =========================
current_row = 16
max_years = 100  # safety buffer

for i in range(max_years):
    ws[f"A{current_row}"] = f"=A{current_row-1}+1"
    ws[f"B{current_row}"] = f"=B{current_row-1}*(1+$B$10)"
    ws[f"C{current_row}"] = f"=C{current_row-1}*(1+$B$11)"
    ws[f"D{current_row}"] = f"=D{current_row-1}*(1+$B$11)"
    ws[f"E{current_row}"] = f"=SUM(B{current_row}:D{current_row})"
    ws[f"F{current_row}"] = f"=F{current_row-1}+E{current_row}"
    current_row += 1

# =========================
# SAVE FILE
# =========================
file_name = "Insurance_Cost_Projection.xlsx"
wb.save(file_name)

print(f"Excel model created: {file_name}")
